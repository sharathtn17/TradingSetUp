import requests
import openpyxl
import re
import time
from pathlib import Path
from datetime import date
import calendar
import smtplib
import random
import json
import nsetools
from nsetools import Nse
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


DZnotRun=[]
today = date.today()
print(calendar.day_name[today.weekday()])
# dd/mm/YY
nse = Nse()
d1 = today.strftime("%Y-%m-%d")
print("d1 =", d1)

xlsx_file = Path('', 'Stockdz.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file) 
sheet = wb_obj.active
  
def triggerMail(mData,sub):
      mail_content = json.dumps(mData)
      #The mail addresses and password
      sender_address = 'sharathoffical73@gmail.com'
      sender_pass = 'PRU$ha2021'
      receiver_address = 'dztrading.21@gmail.com'
      rec_list =  ['dztrading.21@gmail.com']
      rec =  ', '.join(rec_list)
      #Setup the MIME
      message = MIMEMultipart()
      message['From'] = "sharathoffical73@gmail.com"
      message['To'] = rec
      message['Subject'] = sub   #The subject line
      #The body and the attachments for the mail
      message.attach(MIMEText(mail_content, 'plain'))
      #Create SMTP session for sending the mail
      # open the file to be sent 
      #filename = "WIT.xlsx"
      #attachment = open("WIT.xlsx", "rb")

      # instance of MIMEBase and named as p
      p = MIMEBase('application', 'octet-stream')
  
      #   To change the payload into encoded form
      #p.set_payload((attachment).read())
  
      # encode into base64
     # encoders.encode_base64(p)
   
      #p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
  
      # attach the instance 'p' to instance 'msg'
      #message.attach(p)
  
      session = smtplib.SMTP('smtp.gmail.com', 587) #use gmail with port
      session.starttls() #enable security
      session.login(sender_address, sender_pass) #login with mail_id and password
      text = message.as_string()
      session.sendmail(sender_address, rec_list, text)
      session.quit()
      print('Mail Sent')

def getNupdateData(eData,row):
    #Printing Stock
    print(eData[0])
    #Getting quote of specific stock from nsetools 
    quote = nse.get_quote(str(eData[0]))
    
    #Converting quote to string to replace ' single quote to double quote to make it json compatible or to convert to json
    quoteToString=str(quote)
   
    #replace single to double quote using reexp
    quoteToString=re.sub("'","\"",quoteToString)
    #replace None to double quote blank using reexp
    quoteToString=re.sub("None","\" \"",quoteToString)
    #replace False to double quote "FALSE" using reexp
    quoteToString=re.sub("False","\"FALSE\"",quoteToString)
    
    print(quoteToString)
    #f = open('Stockdata.json',)
  
    #Converting string to Json format
    resultJson=json.loads(quoteToString)
    time.sleep(2)
   
    #Get ALL elements in json
    daysLow=resultJson.get('dayLow')
    cmp=resultJson.get('lastPrice')
    pricebandLower=resultJson.get('pricebandlower')
    pricebandUpper=resultJson.get('pricebandupper')
    avgPrice=resultJson.get('averagePrice')
    deliveryToTradedQuantity=resultJson.get('deliveryToTradedQuantity')
    totalTradedVolume=resultJson.get('totalTradedVolume')
    #avgPrice=resultJson.get('averagePrice')
   
    #Read from excel
    stockData={}
    stockData["Stock"]=eData[0]
    stockData["DemandZone"]=eData[2]
    stockData["Days Low"]=daysLow
    stockData["CMP"]=cmp
    stockData["pricebandLower"]=pricebandLower
    stockData["pricebandupper"]=pricebandUpper
    stockData["averagePrice"]=avgPrice
    stockData["Method"]=eData[3]
    stockData["deliveryToTradedQuantity"]=deliveryToTradedQuantity
    stockData["totalTradedVolume"]=totalTradedVolume
    
    #writing to excel
    sheet["H"+str(row)]=str(cmp)
    sheet["J"+str(row)]=str(deliveryToTradedQuantity)
    sheet["K"+str(row)]=str(totalTradedVolume)

    return stockData

def checkDemandZone(stockData,row,Msub):
    ''' #Printing Stock
    print(sData[0])
    
    #Getting quote of specific stock from nsetools 
    quote = nse.get_quote(str(sData[0]))
    #Converting quote to string to replace ' single quote to double quote to make it json compatible or to convert to json
    quoteToString=str(quote)
   
    #replace single to double quote using reexp
    quoteToString=re.sub("'","\"",quoteToString)
    #replace None to double quote blank using reexp
    quoteToString=re.sub("None","\" \"",quoteToString)
    #replace False to double quote "FALSE" using reexp
    quoteToString=re.sub("False","\"FALSE\"",quoteToString)
    print(quoteToString)
    #f = open('Stockdata.json',)
  
    #Converting string to Json format
    resultJson=json.loads(quoteToString)
    
    time.sleep(2)
   
    #Get ALL elements in json
    daysLow=resultJson.get('dayLow')
    cmp=resultJson.get('lastPrice')
    pricebandLower=resultJson.get('pricebandlower')
    pricebandUpper=resultJson.get('pricebandupper')
    avgPrice=resultJson.get('averagePrice')
    deliveryToTradedQuantity=resultJson.get('deliveryToTradedQuantity')
    totalTradedVolume=resultJson.get('totalTradedVolume')
    #avgPrice=resultJson.get('averagePrice')
    
    stockData={}
    stockData["Stock"]=sData[0]
    stockData["DemandZone"]=sData[2]
    stockData["Days Low"]=daysLow
    stockData["CMP"]=cmp
    stockData["pricebandLower"]=pricebandLower
    stockData["pricebandupper"]=pricebandUpper
    stockData["averagePrice"]=avgPrice
    stockData["Method"]=sData[3]
    stockData["deliveryToTradedQuantity"]=deliveryToTradedQuantity
    stockData["totalTradedVolume"]=totalTradedVolume
    #writing to excel
    sheet["H"+str(row)]=str(cmp)
    sheet["J"+str(row)]=str(deliveryToTradedQuantity)
    sheet["K"+str(row)]=str(totalTradedVolume)
    
    print("Stock:"+str(sData[0])+" DZ:"+str(sData[2])+" day'sLow:"+str(daysLow)+" CMP:"+str(cmp)) 
    '''
    intdz=float(stockData["DemandZone"])
    if(float(stockData["Days Low"])<=intdz):
     print("DZ EQUAL TO CMP")
     triggerMail(stockData,Msub)
     sheet["I"+str(row)]="Yes"
     
    time.sleep(2) 


e_list=[]
i=2

for row in sheet.iter_rows(max_row=sheet.max_row-1):
    #for cell in row:
       dzData=[]
       dzData.append(sheet.cell(row=i, column=1).value)
       dzData.append(sheet.cell(row=i, column=2).value)
       dzData.append(sheet.cell(row=i, column=3).value)
       dzData.append(sheet.cell(row=i, column=7).value)
       time.sleep(3)
       try:
        Msub="WacthOut :"+str(sheet.cell(row=i, column=1).value)+" has reached Demand Zone"
        execeldata=getNupdateData(dzData,i)
        checkDemandZone(execeldata,i,Msub)
       except Exception as e:
           DZnotRun.append(sheet.cell(row=i, column=1).value)
           print(e)

           pass

       i+=1
 
wb_obj.save("Stockdz.xlsx")
wb_obj.close()

time.sleep(3)
#open file for reading
xlsx_file = Path('', 'Stockdz.xlsx')
wb_obj1 = openpyxl.load_workbook(xlsx_file) 
sheet2 = wb_obj1.active

k=2
for row in sheet2.iter_rows(max_row=sheet2.max_row-1):
    #for cell in row:
       upsideData={}
       upsideData["Stock"]=sheet2.cell(row=k, column=1).value
       upsideData["DZ"]=sheet2.cell(row=k, column=3).value
       upsideData["CMP"]=sheet2.cell(row=k, column=8).value
       upsideData["TouchedDZ"]=sheet2.cell(row=k, column=9).value
       print("#####Upside#######")
       print(str(upsideData["DZ"]))
       print(str(upsideData["CMP"]))  
       print(str(upsideData["TouchedDZ"]))  
    
       if(str(upsideData["TouchedDZ"])=="Yes"):
         up=(float(upsideData["CMP"])-float(upsideData["DZ"]))
         print(str(up))
         down=float(upsideData["DZ"])
         p1=up/down
         print(p1)
         percentage=p1*100
         print(percentage)
         if(percentage >=3):
            upsideData["PercentageUP: "]=round(percentage)
            Msub='Alert :'+str(upsideData["Stock"])+": Moving more than 3 percentage from Demand Zone"
            triggerMail(upsideData,Msub)
            print("#####HURRAY#######"+str(percentage))
           
       

       k+=1


j=0
if(len(DZnotRun)!=0):
     emap={}
     emap["DZ"]=DZnotRun
     triggerMail(emap,"Stocks DZ not RUN")
    